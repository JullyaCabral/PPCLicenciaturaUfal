"""
Módulo de exportações para componentes curriculares.
Responsável por gerar arquivos CSV, XLSX e PDF.
"""

import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT


def exportar_csv(componentes: list, caminho_arquivo: str, tabela: str = "componentes") -> str:
    """
    Exporta dados para CSV, permitindo escolher qual tabela será gerada.
    
    Args:
        componentes: Lista de dicionários com os componentes
        caminho_arquivo: Caminho onde o arquivo será salvo
        tabela: Nome da tabela desejada (componentes, matriz, resumo_nucleo)
    
    Returns:
        Caminho do arquivo salvo
    """
    tabela_normalizada = (tabela or "componentes").lower()
    
    if tabela_normalizada == "matriz":
        df = gerar_matriz_por_periodo(componentes)
    elif tabela_normalizada in {"resumo", "resumo_nucleo", "por_nucleo"}:
        df = gerar_resumo_por_semestre_nucleo(componentes)
    else:
        dados_csv = []
        for comp in componentes:
            linha = {
                "Semestre": comp.get("semestre", ""),
                "Nome": comp.get("nome", ""),
                "Tipo": comp.get("tipo", ""),
                "Aulas Semanais": comp.get("aulas_semanais", ""),
                "CH Total": comp.get("ch_total", ""),
                "CH Teórica": comp.get("ch_teorica", ""),
                "CH Prática": comp.get("ch_pratica", ""),
                "CH Extensão": comp.get("ch_extensao", ""),
                "Núcleo": comp.get("nucleo", ""),
                "Temas Núcleo I": "; ".join(comp.get("temas_nucleo_i", [])) if comp.get("temas_nucleo_i") else "",
                "Diretrizes Núcleo II": comp.get("diretrizes_nucleo_ii", ""),
                "Descrição Extensão": comp.get("descricao_extensao", ""),
                "Local Realização": comp.get("local_realizacao", ""),
                "Etapa Estágio": comp.get("etapa_estagio", ""),
                "Bloco": comp.get("bloco", ""),
                "Observações": comp.get("observacoes", "")
            }
            dados_csv.append(linha)
        df = pd.DataFrame(dados_csv)
    
    df.to_csv(caminho_arquivo, index=False, encoding="utf-8-sig", sep=";")
    
    return caminho_arquivo


def exportar_xlsx(componentes: list, caminho_arquivo: str, abas: list[str] | None = None) -> str:
    """
    Exporta dados para arquivo XLSX, permitindo selecionar quais abas devem ser geradas.
    
    Args:
        componentes: Lista de dicionários com os componentes
        caminho_arquivo: Caminho onde o arquivo será salvo
        abas: Lista de abas desejadas (matriz, resumo_nucleo, componentes)
    
    Returns:
        Caminho do arquivo salvo
    """
    from openpyxl.utils import get_column_letter
    
    abas_padrao = ["matriz", "resumo_nucleo", "componentes"]
    abas_normalizadas = [aba.lower() for aba in (abas or abas_padrao) if aba]
    
    if not abas_normalizadas:
        raise ValueError("Selecione ao menos uma aba para exportação.")
    
    with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
        if "matriz" in abas_normalizadas:
            df_matriz = gerar_matriz_por_periodo(componentes)
            df_matriz.to_excel(writer, sheet_name="Matriz", index=False)
            worksheet_matriz = writer.sheets["Matriz"]
            for idx, col in enumerate(df_matriz.columns, 1):
                max_length = max(
                    df_matriz[col].astype(str).apply(len).max() if len(df_matriz) > 0 else 0,
                    len(col)
                )
                col_letter = get_column_letter(idx)
                worksheet_matriz.column_dimensions[col_letter].width = min(max_length + 2, 50)
        
        if "resumo_nucleo" in abas_normalizadas or "por_nucleo" in abas_normalizadas:
            df_nucleo = gerar_resumo_por_semestre_nucleo(componentes)
            df_nucleo.to_excel(writer, sheet_name="Por Núcleo", index=False)
            worksheet_nucleo = writer.sheets["Por Núcleo"]
            for idx, col in enumerate(df_nucleo.columns, 1):
                max_length = max(
                    df_nucleo[col].astype(str).apply(len).max() if len(df_nucleo) > 0 else 0,
                    len(col)
                )
                col_letter = get_column_letter(idx)
                worksheet_nucleo.column_dimensions[col_letter].width = min(max_length + 2, 50)
        
        if "componentes" in abas_normalizadas:
            dados_componentes = []
            for comp in componentes:
                linha = {
                    "Semestre": comp.get("semestre", ""),
                    "Nome": comp.get("nome", ""),
                    "Tipo": comp.get("tipo", ""),
                    "Aulas Semanais": comp.get("aulas_semanais", ""),
                    "CH Total": comp.get("ch_total", ""),
                    "CH Teórica": comp.get("ch_teorica", ""),
                    "CH Prática": comp.get("ch_pratica", ""),
                    "CH Extensão": comp.get("ch_extensao", ""),
                    "Núcleo": comp.get("nucleo", ""),
                    "Temas Núcleo I": "; ".join(comp.get("temas_nucleo_i", [])) if comp.get("temas_nucleo_i") else "",
                    "Diretrizes Núcleo II": comp.get("diretrizes_nucleo_ii", ""),
                    "Descrição Extensão": comp.get("descricao_extensao", ""),
                    "Local Realização": comp.get("local_realizacao", ""),
                    "Etapa Estágio": comp.get("etapa_estagio", ""),
                    "Bloco": comp.get("bloco", ""),
                    "Observações": comp.get("observacoes", "")
                }
                dados_componentes.append(linha)
            
            df_componentes = pd.DataFrame(dados_componentes)
            df_componentes.to_excel(writer, sheet_name="Componentes", index=False)
            worksheet_comp = writer.sheets["Componentes"]
            for idx, col in enumerate(df_componentes.columns, 1):
                max_length = max(
                    df_componentes[col].astype(str).apply(len).max() if len(df_componentes) > 0 else 0,
                    len(col)
                )
                col_letter = get_column_letter(idx)
                worksheet_comp.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    return caminho_arquivo


def _ordenar_semestre_valor(semestre) -> tuple:
    if isinstance(semestre, (int, float)):
        return (0, int(semestre))
    if isinstance(semestre, str):
        sem_strip = semestre.strip()
        if sem_strip.isdigit():
            return (0, int(sem_strip))
        return (1, sem_strip.lower())
    return (1, "")


def _formatar_rotulo_periodo(semestre) -> str:
    if isinstance(semestre, (int, float)):
        return f"{int(semestre)}º Período"
    if isinstance(semestre, str) and semestre.strip():
        return semestre.strip()
    return "Sem período"


def _obter_observacao_nucleo(componente: dict) -> str:
    nucleo = componente.get("nucleo")
    partes: list[str] = []
    
    if nucleo == "I":
        temas = componente.get("temas_nucleo_i") or []
        codigos = []
        for tema in temas:
            if isinstance(tema, str) and ")" in tema:
                codigos.append(tema.split(")")[0].strip().upper())
            elif isinstance(tema, str) and tema.strip():
                codigos.append(tema.strip().upper())
        if codigos:
            partes.append(f"TEMA: {', '.join(codigos)}")
    
    elif nucleo == "II":
        diretrizes = (componente.get("diretrizes_nucleo_ii") or "").strip()
        if diretrizes:
            partes.append(f"Diretrizes: {diretrizes}")
    
    elif nucleo == "III":
        descricao = (componente.get("descricao_extensao") or "").strip()
        if descricao:
            partes.append(f"Extensão: {descricao}")
    
    elif nucleo == "IV":
        local = (componente.get("local_realizacao") or "").strip()
        etapa = (componente.get("etapa_estagio") or "").strip()
        detalhes = []
        if local:
            detalhes.append(f"Local: {local}")
        if etapa:
            detalhes.append(f"Etapa: {etapa}")
        if detalhes:
            partes.append(" | ".join(detalhes))
    
    observacao_geral = (componente.get("observacoes") or "").strip()
    if observacao_geral:
        partes.append(f"Obs.: {observacao_geral}")
    
    return " | ".join(partes)


def gerar_matriz_por_periodo(componentes: list) -> pd.DataFrame:
    """
    Gera a matriz curricular principal organizada por período/semestre.
    Inclui linha de cabeçalho por período e linha TOTAL por período.
    """
    colunas = [
        "Semestre",
        "Nome",
        "Tipo",
        "CH Semanal",
        "CH Teórica",
        "CH Prática",
        "CH Extensão",
        "CH Total",
        "Núcleo",
        "Observação Núcleo"
    ]
    
    if not componentes:
        return pd.DataFrame(columns=colunas)
    
    componentes_ordenados = sorted(
        componentes,
        key=lambda x: (
            _ordenar_semestre_valor(x.get("semestre")),
            (x.get("nome") or "").lower()
        )
    )
    
    grupos: dict = {}
    ordem_grupos: list = []
    for comp in componentes_ordenados:
        semestre_val = comp.get("semestre")
        if isinstance(semestre_val, (int, float)) and semestre_val > 0:
            chave = int(semestre_val)
        elif isinstance(semestre_val, str) and semestre_val.strip().isdigit():
            chave = int(semestre_val.strip())
        else:
            chave = "Sem período"
        
        if chave not in grupos:
            grupos[chave] = []
            ordem_grupos.append(chave)
        grupos[chave].append(comp)
    
    dados_matriz: list[dict] = []
    
    for chave in ordem_grupos:
        componentes_semestre = grupos[chave]
        rotulo_periodo = _formatar_rotulo_periodo(chave)
        dados_matriz.append({
            "Semestre": rotulo_periodo,
            "Nome": "",
            "Tipo": "",
            "CH Semanal": None,
            "CH Teórica": None,
            "CH Prática": None,
            "CH Extensão": None,
            "CH Total": None,
            "Núcleo": "",
            "Observação Núcleo": ""
        })
        
        for comp in componentes_semestre:
            aulas_semanais_val = comp.get("aulas_semanais")
            if isinstance(aulas_semanais_val, (int, float)):
                aulas_semanais_display = int(aulas_semanais_val)
            else:
                aulas_semanais_display = None
            
        semestre_bruto = comp.get("semestre", "")
        if isinstance(semestre_bruto, (int, float)):
            semestre_formatado = str(int(semestre_bruto))
        else:
            semestre_formatado = str(semestre_bruto) if semestre_bruto not in (None, "") else ""
        
        dados_matriz.append({
            "Semestre": semestre_formatado,
                "Nome": comp.get("nome", ""),
                "Tipo": comp.get("tipo", ""),
                "CH Semanal": aulas_semanais_display,
                "CH Teórica": comp.get("ch_teorica", 0),
                "CH Prática": comp.get("ch_pratica", 0),
                "CH Extensão": comp.get("ch_extensao", 0),
                "CH Total": comp.get("ch_total", 0),
                "Núcleo": comp.get("nucleo", ""),
                "Observação Núcleo": _obter_observacao_nucleo(comp)
            })
        
        ch_total_sem = sum(c.get("ch_total", 0) for c in componentes_semestre)
        ch_teorica_sem = sum(c.get("ch_teorica", 0) for c in componentes_semestre)
        ch_pratica_sem = sum(c.get("ch_pratica", 0) for c in componentes_semestre)
        ch_extensao_sem = sum(c.get("ch_extensao", 0) for c in componentes_semestre)
        
        dados_matriz.append({
            "Semestre": str(chave),
            "Nome": "TOTAL DO PERÍODO",
            "Tipo": "",
            "CH Semanal": None,
            "CH Teórica": ch_teorica_sem,
            "CH Prática": ch_pratica_sem,
            "CH Extensão": ch_extensao_sem,
            "CH Total": ch_total_sem,
            "Núcleo": "",
            "Observação Núcleo": ""
        })
    
    return pd.DataFrame(dados_matriz, columns=colunas)


def gerar_resumo_por_semestre_nucleo(componentes: list) -> pd.DataFrame:
    """
    Gera um resumo da carga horária por semestre e núcleo.
    
    Args:
        componentes: Lista de dicionários com os componentes
    
    Returns:
        DataFrame com resumo por semestre e núcleo
    """
    from utils.calculos import calcular_ch_por_nucleo
    
    # Obter todos os semestres únicos
    semestres = sorted(set(comp.get("semestre", 0) for comp in componentes if comp.get("semestre")))
    
    dados_resumo = []
    
    for semestre in semestres:
        # Filtrar componentes do semestre
        comps_semestre = [c for c in componentes if c.get("semestre") == semestre]
        
        # Calcular CH por núcleo neste semestre
        ch_i = sum(c.get("ch_total", 0) for c in comps_semestre if c.get("nucleo") == "I")
        ch_ii = sum(c.get("ch_total", 0) for c in comps_semestre if c.get("nucleo") == "II")
        ch_iii = sum(c.get("ch_total", 0) for c in comps_semestre if c.get("nucleo") == "III")
        ch_iv = sum(c.get("ch_total", 0) for c in comps_semestre if c.get("nucleo") == "IV")
        ch_total_semestre = ch_i + ch_ii + ch_iii + ch_iv
        
        dados_resumo.append({
            "Semestre": str(semestre),  # Converter para string para evitar conflito de tipos
            "CH Núc. I": ch_i,
            "CH Núc. II": ch_ii,
            "CH Núc. III": ch_iii,
            "CH Núc. IV": ch_iv,
            "Total": ch_total_semestre
        })
    
    # Linha de totais
    if dados_resumo:
        total_i = sum(r["CH Núc. I"] for r in dados_resumo)
        total_ii = sum(r["CH Núc. II"] for r in dados_resumo)
        total_iii = sum(r["CH Núc. III"] for r in dados_resumo)
        total_iv = sum(r["CH Núc. IV"] for r in dados_resumo)
        total_geral = sum(r["Total"] for r in dados_resumo)
        
        dados_resumo.append({
            "Semestre": "TOTAL",
            "CH Núc. I": total_i,
            "CH Núc. II": total_ii,
            "CH Núc. III": total_iii,
            "CH Núc. IV": total_iv,
            "Total": total_geral
        })
    
    return pd.DataFrame(dados_resumo)


def _agrupar_componentes_por_semestre(componentes: list) -> list[dict]:
    """
    Agrupa componentes por semestre, ordenando e calculando totais auxiliares.
    """
    grupos: dict[int | str, list[dict]] = {}
    
    for comp in componentes:
        semestre_val = comp.get("semestre")
        if isinstance(semestre_val, (int, float)) and semestre_val > 0:
            chave = int(semestre_val)
        else:
            chave = "Sem período"
        grupos.setdefault(chave, []).append(comp)
    
    resultado = []
    for chave in sorted(grupos, key=_ordenar_semestre_valor):
        componentes_semestre = sorted(grupos[chave], key=lambda x: (x.get("nome") or "").lower())
        ch_total = sum(c.get("ch_total", 0) for c in componentes_semestre)
        ch_teorica = sum(c.get("ch_teorica", 0) for c in componentes_semestre)
        ch_pratica = sum(c.get("ch_pratica", 0) for c in componentes_semestre)
        ch_extensao = sum(c.get("ch_extensao", 0) for c in componentes_semestre)
        resultado.append({
            "rotulo": _formatar_rotulo_periodo(chave),
            "componentes": componentes_semestre,
            "totais": {
                "ch_total": ch_total,
                "ch_teorica": ch_teorica,
                "ch_pratica": ch_pratica,
                "ch_extensao": ch_extensao
            }
        })
    return resultado


def _formatar_carga_horaria(valor: float | int | str | None) -> str:
    if valor in (None, "", 0):
        return ""
    try:
        return f"{float(valor):.0f}h"
    except (ValueError, TypeError):
        return str(valor)


def _formatar_aulas_semanais(valor: float | int | None) -> str:
    if isinstance(valor, (int, float)) and valor > 0:
        return f"{int(valor)}"
    return ""


def _formatar_celula_matriz_pdf(valor, coluna: str) -> str:
    if valor is None:
        return ""
    if isinstance(valor, str):
        return valor
    try:
        if isinstance(valor, (int, float)):
            if float(valor).is_integer():
                return str(int(valor))
            return f"{valor:.2f}"
    except (ValueError, TypeError):
        return str(valor)
    return str(valor)


def exportar_pdf(componentes: list, caminho_arquivo: str, secoes: list[str] | None = None) -> str:
    """
    Exporta um relatório em PDF configurável, com possibilidade de escolher seções.
    
    Args:
        componentes: Lista de dicionários com os componentes
        caminho_arquivo: Caminho onde o arquivo será salvo
        secoes: Lista de seções desejadas (matriz, resumo_nucleo, resumo_geral, conformidade)
    
    Returns:
        Caminho do arquivo salvo
    """
    from utils.calculos import (
        calcular_ch_total_curso,
        calcular_ch_por_nucleo,
        calcular_percentual_extensao,
        calcular_percentual_pratica_pedagogica,
        obter_ch_minima_por_nucleo,
        validar_ch_minima_nucleo
    )
    from utils.validacoes import validar_curso_completo
    
    secoes_padrao = ["matriz", "resumo_nucleo", "resumo_geral", "conformidade"]
    secoes_normalizadas = [sec.lower() for sec in (secoes or secoes_padrao) if sec]
    
    if not secoes_normalizadas:
        raise ValueError("Selecione ao menos uma seção para exportação.")
    
    doc = SimpleDocTemplate(
        caminho_arquivo,
        pagesize=A4,
        leftMargin=1.2 * cm,
        rightMargin=1.0 * cm,
        topMargin=1.6 * cm,
        bottomMargin=1.4 * cm
    )
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#0B5FA5'),
        spaceAfter=18,
        alignment=TA_CENTER
    )
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#0B5FA5'),
        spaceAfter=10,
        leading=14
    )
    table_text_style = ParagraphStyle(
        'TableText',
        parent=styles['Normal'],
        fontSize=7,
        leading=9,
        alignment=TA_LEFT
    )
    table_header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontSize=7,
        leading=9,
        alignment=TA_CENTER,
        textColor=colors.whitesmoke
    )
    
    story.append(Paragraph("Relatório de Carga Horária - Componentes Curriculares", title_style))
    story.append(Spacer(1, 0.25 * cm))
    
    if "matriz" in secoes_normalizadas:
        story.append(Paragraph("Matriz Curricular por Período", heading_style))
        blocos = _agrupar_componentes_por_semestre(componentes)
        if not blocos:
            story.append(Paragraph("Nenhum componente cadastrado.", table_text_style))
        else:
            colunas_pdf = [
                "Nome do Componente",
                "Tipo",
                "CH Semanal",
                "CH Teórica",
                "CH Prática",
                "CH Extensão",
                "CH Total",
                "Núcleo",
                "Observação do Núcleo"
            ]
            dados_tabela = [[Paragraph(titulo, table_header_style) for titulo in colunas_pdf]]
            estilos_especificos: list[tuple] = []
            
            for bloco in blocos:
                linha_periodo_idx = len(dados_tabela)
                linha_periodo = [Paragraph(f"{bloco['rotulo']}", table_header_style)] + [""] * (len(colunas_pdf) - 1)
                dados_tabela.append(linha_periodo)
                estilos_especificos.extend([
                    ('SPAN', (0, linha_periodo_idx), (-1, linha_periodo_idx)),
                    ('BACKGROUND', (0, linha_periodo_idx), (-1, linha_periodo_idx), colors.HexColor('#E8EFF9')),
                    ('FONTNAME', (0, linha_periodo_idx), (-1, linha_periodo_idx), 'Helvetica-Bold')
                ])
                
                for comp in bloco["componentes"]:
                    linha = [
                        Paragraph(comp.get("nome", "") or "-", table_text_style),
                        comp.get("tipo", "") or "",
                        _formatar_celula_matriz_pdf(_formatar_aulas_semanais(comp.get("aulas_semanais")), "CH Semanal"),
                        _formatar_celula_matriz_pdf(comp.get("ch_teorica", 0), "CH Teórica"),
                        _formatar_celula_matriz_pdf(comp.get("ch_pratica", 0), "CH Prática"),
                        _formatar_celula_matriz_pdf(comp.get("ch_extensao", 0), "CH Extensão"),
                        _formatar_celula_matriz_pdf(comp.get("ch_total", 0), "CH Total"),
                        comp.get("nucleo", "") or "",
                        Paragraph(_obter_observacao_nucleo(comp), table_text_style)
                    ]
                    dados_tabela.append(linha)
                
                totais = bloco["totais"]
                linha_total_idx = len(dados_tabela)
                dados_tabela.append([
                    Paragraph("<b>TOTAL DO PERÍODO</b>", table_text_style),
                    "",
                    "",
                    _formatar_celula_matriz_pdf(totais["ch_teorica"], "CH Teórica"),
                    _formatar_celula_matriz_pdf(totais["ch_pratica"], "CH Prática"),
                    _formatar_celula_matriz_pdf(totais["ch_extensao"], "CH Extensão"),
                    _formatar_celula_matriz_pdf(totais["ch_total"], "CH Total"),
                    "",
                    ""
                ])
                estilos_especificos.append(('FONTNAME', (0, linha_total_idx), (-1, linha_total_idx), 'Helvetica-Bold'))
                estilos_especificos.append(('BACKGROUND', (0, linha_total_idx), (-1, linha_total_idx), colors.HexColor('#F2F5FA')))
            
            col_widths = [
                5.9 * cm,  # Nome
                1.85 * cm,  # Tipo
                1.4 * cm,  # CH Semanal
                1.25 * cm,  # CH Teórica
                1.25 * cm,  # CH Prática
                1.4 * cm,  # CH Extensão
                1.25 * cm,  # CH Total
                1.25 * cm,  # Núcleo
                2.70 * cm   # Observação
            ]
            
            tabela_matriz = Table(dados_tabela, repeatRows=1, colWidths=col_widths)
            estilo_base = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0B5FA5')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                ('ALIGN', (1, 1), (-3, -1), 'CENTER'),
                ('ALIGN', (-2, 1), (-2, -1), 'CENTER'),
                ('ALIGN', (-1, 1), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#B5C6E0')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F6F8FC')]),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]
            tabela_matriz.setStyle(TableStyle(estilo_base + estilos_especificos))
            story.append(tabela_matriz)
            story.append(Spacer(1, 0.35 * cm))
    
    if "resumo_nucleo" in secoes_normalizadas or "por_nucleo" in secoes_normalizadas:
        story.append(Paragraph("Quadro-Resumo: CH por Semestre e Núcleo", heading_style))
        df_resumo = gerar_resumo_por_semestre_nucleo(componentes)
        dados_resumo = [list(df_resumo.columns)]
        for _, row in df_resumo.iterrows():
            dados_resumo.append([
                str(row["Semestre"]),
                _formatar_carga_horaria(row["CH Núc. I"]),
                _formatar_carga_horaria(row["CH Núc. II"]),
                _formatar_carga_horaria(row["CH Núc. III"]),
                _formatar_carga_horaria(row["CH Núc. IV"]),
                _formatar_carga_horaria(row["Total"])
            ])
        
        tabela_resumo_nucleo = Table(
            dados_resumo,
            repeatRows=1,
            colWidths=[2.5 * cm, 2.6 * cm, 2.6 * cm, 2.6 * cm, 2.6 * cm, 2.6 * cm]
        )
        tabela_resumo_nucleo.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0B5FA5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F6F8FC')]),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E5EDF9')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#B5C6E0')),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(tabela_resumo_nucleo)
        story.append(Spacer(1, 0.35 * cm))
    
    if "resumo_geral" in secoes_normalizadas:
        story.append(Paragraph("Resumo Geral do Curso", heading_style))
        ch_total = calcular_ch_total_curso(componentes)
        ch_i = calcular_ch_por_nucleo(componentes, "I")
        ch_ii = calcular_ch_por_nucleo(componentes, "II")
        ch_iii = calcular_ch_por_nucleo(componentes, "III")
        ch_iv = calcular_ch_por_nucleo(componentes, "IV")
        perc_extensao = calcular_percentual_extensao(componentes)
        perc_pratica = calcular_percentual_pratica_pedagogica(componentes)
        
        resumo_geral = [
            ["Carga Horária Total do Curso", _formatar_carga_horaria(ch_total)],
            ["CH Núcleo I", _formatar_carga_horaria(ch_i)],
            ["CH Núcleo II", _formatar_carga_horaria(ch_ii)],
            ["CH Núcleo III", _formatar_carga_horaria(ch_iii)],
            ["CH Núcleo IV", _formatar_carga_horaria(ch_iv)],
            ["Percentual de Extensão", f"{perc_extensao:.2f}%"],
            ["Percentual de Prática Pedagógica", f"{perc_pratica:.2f}%"]
        ]
        
        tabela_resumo = Table(resumo_geral, colWidths=[9.0 * cm, 5.0 * cm])
        tabela_resumo.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0B5FA5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#B5C6E0')),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(tabela_resumo)
        story.append(Spacer(1, 0.35 * cm))
    else:
        # Necessário para cálculo de conformidade
        ch_total = calcular_ch_total_curso(componentes)
        ch_i = calcular_ch_por_nucleo(componentes, "I")
        ch_ii = calcular_ch_por_nucleo(componentes, "II")
        ch_iii = calcular_ch_por_nucleo(componentes, "III")
        ch_iv = calcular_ch_por_nucleo(componentes, "IV")
        perc_extensao = calcular_percentual_extensao(componentes)
    
    if "conformidade" in secoes_normalizadas:
        story.append(Paragraph("Resumo de Conformidade", heading_style))
        resultado_validacao = validar_curso_completo(componentes)
        conformidade_itens = []
        
        for nucleo in ["I", "II", "III", "IV"]:
            ch_atual = calcular_ch_por_nucleo(componentes, nucleo)
            ch_minima = obter_ch_minima_por_nucleo(nucleo)
            valido, _ = validar_ch_minima_nucleo(ch_atual, ch_minima)
            status = "✓" if valido else "✗"
            conformidade_itens.append([f"{status} Núcleo {nucleo} (mín. {ch_minima:.0f}h)", _formatar_carga_horaria(ch_atual)])
        
        status_total = "✓" if ch_total >= 3200 else "✗"
        status_ext = "✓" if perc_extensao >= 10 else "✗"
        
        conformidade_itens.extend([
            [f"{status_total} CH Total do Curso (mín. 3200h)", _formatar_carga_horaria(ch_total)],
            [f"{status_ext} Percentual de Extensão (mín. 10%)", f"{perc_extensao:.2f}%"]
        ])
        
        if resultado_validacao["erros"]:
            conformidade_itens.append([f"✗ Pendências detectadas ({len(resultado_validacao['erros'])})", ""])
        else:
            conformidade_itens.append(["✓ Curso conforme com todas as validações principais", ""])
        
        tabela_conformidade = Table(conformidade_itens, colWidths=[9.0 * cm, 5.0 * cm])
        tabela_conformidade.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0B5FA5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#B5C6E0')),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(tabela_conformidade)
    
    doc.build(story)
    return caminho_arquivo

